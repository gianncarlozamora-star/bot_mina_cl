"""
EXPORTACIÓN A EXCEL
Genera el reporte de avance diario en formato Excel
compatible con el AVANCE_DIARIO_DDH.xlsx existente.
"""
import io
from datetime import datetime
from db.conexion import ejecutar
from config import hora_peru

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


def generar_avance_diario(mes: int = None, anio: int = None) -> bytes | None:
    """
    Genera el Excel de avance diario.
    Si no se especifica mes/año, usa el mes actual.
    Retorna bytes del archivo o None si falla.
    """
    if not OPENPYXL_OK:
        return None

    hoy   = hora_peru()
    mes   = mes  or hoy.month
    anio  = anio or hoy.year

    rows = ejecutar(
        """SELECT
               EXTRACT(YEAR  FROM ap.fecha)::INTEGER,
               EXTRACT(MONTH FROM ap.fecha)::INTEGER,
               ap.fecha,
               ap.turno,
               e.codigo,
               m.codigo,
               sc.nombre,
               s.nivel_prog,
               s.bhid,
               ap.prof_inicio,
               ap.prof_final,
               ap.metros_avance,
               ap.budget_metros,
               ap.valor_usd,
               ap.valor_budget,
               s.cuerpo_objetivo,
               s.tajo_objetivo,
               ap.observaciones
           FROM avance_perforacion ap
           JOIN sondajes          s  ON ap.sondaje_id   = s.id
           JOIN cat_maquinas      m  ON ap.maquina_id   = m.id
           JOIN cat_empresas      e  ON m.empresa_id    = e.id
           JOIN cat_subcategorias sc ON s.subcategoria_id = sc.id
           WHERE EXTRACT(MONTH FROM ap.fecha) = %s
             AND EXTRACT(YEAR  FROM ap.fecha) = %s
             AND ap.estado = 'ACTIVO'
           ORDER BY ap.fecha, e.codigo, m.codigo""",
        (mes, anio), fetchall=True
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # ── ESTILOS ───────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    MESES_ES = {1:"ENERO",2:"FEBRERO",3:"MARZO",4:"ABRIL",5:"MAYO",6:"JUNIO",
                7:"JULIO",8:"AGOSTO",9:"SETIEMBRE",10:"OCTUBRE",
                11:"NOVIEMBRE",12:"DICIEMBRE"}

    # ── ENCABEZADOS ───────────────────────────────────────────
    headers = [
        "AÑO","MES","FECHA","GUARDIA","E.E.","MAQUINA",
        "CATEGORIA","NIVEL","TALADRO","DESDE","HASTA","METROS",
        "BUDGET","LINEA","VALOR $","VALOR $ BUDGET",
        "OBJETIVOS","NOMBRE TAJO","OBSERVACIONES"
    ]
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(header)+4, 12)

    # ── DATOS ─────────────────────────────────────────────────
    fill_par  = PatternFill("solid", start_color="EBF3FB")
    fill_impar= PatternFill("solid", start_color="FFFFFF")

    for row_num, row in enumerate(rows, 2):
        anio_r, mes_r, fecha, turno, empresa, maquina, categoria, nivel, \
        bhid, desde, hasta, metros, budget, valor, vbudget, obj, tajo, obs = row

        fill = fill_par if row_num % 2 == 0 else fill_impar

        # Calcular línea tarifaria
        linea = _calcular_linea(bhid, desde, hasta, empresa)

        ws_row = [
            anio_r,
            MESES_ES.get(mes_r, ""),
            fecha.strftime("%d/%m/%Y") if hasattr(fecha, "strftime") else str(fecha),
            turno,
            empresa, maquina, categoria, nivel, bhid,
            float(desde or 0), float(hasta or 0), float(metros or 0),
            float(budget or 0), linea,
            float(valor or 0), float(vbudget or 0),
            obj or "", tajo or "", obs or ""
        ]
        ws.append(ws_row)

        for col_num in range(1, len(ws_row)+1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill   = fill
            cell.border = thin_border
            if col_num in (10,11,12,13,15,16):  # numéricos
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_num in (1,2,4,5,6,8):  # centrados
                cell.alignment = center_align

    # ── FILA DE TOTALES ───────────────────────────────────────
    total_row = ws.max_row + 1
    ws.cell(total_row, 1, "TOTAL").font = Font(bold=True)
    for col in [12, 13, 15, 16]:
        col_letter = get_column_letter(col)
        ws.cell(total_row, col, f"=SUM({col_letter}2:{col_letter}{total_row-1})")
        ws.cell(total_row, col).font         = Font(bold=True)
        ws.cell(total_row, col).number_format= "#,##0.00"

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # ── GUARDAR EN MEMORIA ────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def generar_estado_sondajes() -> bytes | None:
    """Genera Excel con el estado actual de todos los sondajes."""
    if not OPENPYXL_OK:
        return None

    rows = ejecutar(
        """SELECT s.bhid, sc.nombre, s.tajo_objetivo, s.cuerpo_objetivo,
                  s.campana, m.codigo, e.codigo,
                  s.profundidad_prog, s.profundidad_final,
                  s.estado_logueo, s.estado_muestreo, s.estado_rqd,
                  s.estado_fotografia, s.estado_densidad,
                  s.estado_laboratorio, s.estado_modelado,
                  s.estado_desviacion, s.nivel_prog, s.labor, s.diametro
           FROM sondajes s
           JOIN cat_subcategorias sc ON s.subcategoria_id = sc.id
           JOIN cat_maquinas      m  ON s.maquina_id      = m.id
           JOIN cat_empresas      e  ON s.empresa_id      = e.id
           WHERE sc.fase_activa = TRUE
           ORDER BY s.bhid""",
        fetchall=True
    )

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Estado_DDH"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E79")

    headers = [
        "BHID","SUBCATEGORIA","TAJO","CUERPO","CAMPAÑA",
        "MÁQUINA","E.E.","PROG (m)","FINAL (m)",
        "LOGUEO","MUESTREO","RQD","FOTOGRAFÍA","DENSIDAD",
        "LABORATORIO","MODELADO","DESVIACIÓN","NIVEL","LABOR","DIÁMETRO"
    ]
    ws.append(headers)
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(1, col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_num)].width = max(len(h)+3, 10)

    # Colores por estado
    colores = {
        "COMPLETADO": "C6EFCE",
        "EN_PROCESO": "FFEB9C",
        "PENDIENTE":  "FFCCCC",
        "NO_APLICA":  "D9D9D9",
    }

    for row_num, row in enumerate(rows, 2):
        ws.append(list(row))
        # Colorear columnas de estado (10-16)
        for col in range(10, 17):
            val  = str(row[col-1] or "")
            fill_color = colores.get(val, "FFFFFF")
            ws.cell(row_num, col).fill = PatternFill("solid", start_color=fill_color)

    ws.freeze_panes = "A2"
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _calcular_linea(bhid: str, desde: float, hasta: float, empresa: str) -> str:
    """Determina la línea tarifaria principal del avance."""
    if not bhid:
        return ""
    row = ejecutar(
        "SELECT diametro FROM sondajes WHERE bhid=%s", (bhid,), fetchone=True
    )
    if not row:
        return ""
    diametro = row[0]
    metro_mid = (float(desde or 0) + float(hasta or 0)) / 2
    tramo_desde = int(metro_mid / 100) * 100
    tramo_hasta = tramo_desde + 100
    sufijo = "ED" if empresa == "EXPLODRILLING" else ""
    return f"{diametro}({tramo_desde}-{tramo_hasta}){' '+sufijo if sufijo else ''}"
